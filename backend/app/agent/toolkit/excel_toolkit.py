# ========= Copyright 2025-2026 @ Eigent.ai All Rights Reserved. =========
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
# ========= Copyright 2025-2026 @ Eigent.ai All Rights Reserved. =========

import os
from camel.toolkits import ExcelToolkit as BaseExcelToolkit

from app.component.environment import env
from app.service.task import Agents
from app.utils.listen.toolkit_listen import auto_listen_toolkit, listen_toolkit
from app.utils.toolkit.abstract_toolkit import AbstractToolkit
import logging

logger = logging.getLogger("excel_toolkit")


@auto_listen_toolkit(BaseExcelToolkit)
class ExcelToolkit(BaseExcelToolkit, AbstractToolkit):
    agent_name: str = Agents.document_agent

    def __init__(
        self,
        api_task_id: str,
        timeout: float | None = None,
        working_directory: str | None = None,
    ):
        self.api_task_id = api_task_id
        if working_directory is None:
            working_directory = env("EIGENT_DATA_DIR", os.path.expanduser("~/.eigent/server_data"))
        super().__init__(timeout=timeout, working_directory=working_directory)

    @listen_toolkit(
        inputs=lambda _, document_path, include_cell_info: f"extract the content of the Excel file: {document_path}, {'with metadata' if include_cell_info else 'without metadata'}",
        return_msg=lambda res: f"Excel file content extracted with {len(res)} characters",
    )
    def extract_excel_content(self, document_path: str, include_cell_info: bool = False) -> str:
        r"""Extract and analyze the full content of an Excel file (.xlsx/.xls/.
        csv).

        Use this tool to read and understand the structure and content of
        Excel files. This is typically the first step when working with
        existing Excel files.

        Args:
            document_path (str): The file path to the Excel file.
            include_cell_info (bool): Whether to include detailed per-cell metadata
                (value, font color, fill color, position). Defaults to False.

        Returns:
            str: A comprehensive report containing:
                - Sheet names and their content in markdown table format
                - Optional detailed cell information including values, colors, and
                    positions
                - Formatted data that's easy to understand and analyze
        """
        import pandas as pd
        from openpyxl import load_workbook
        from xls2xlsx import XLS2XLSX

        logger.debug(
            f"Calling extract_excel_content with document_path"
            f": {document_path}"
        )

        if not self._validate_file_path(document_path):
            return "Error: Invalid file path."

        if not (
            document_path.endswith("xls")
            or document_path.endswith("xlsx")
            or document_path.endswith("csv")
        ):
            logger.error("Only xls, xlsx, csv files are supported.")
            return (
                f"Failed to process file {document_path}: "
                f"It is not excel format. Please try other ways."
            )

        if not os.path.exists(document_path):
            return f"Error: File {document_path} does not exist."

        if document_path.endswith("csv"):
            try:
                df = pd.read_csv(document_path)
                md_table = self._convert_to_markdown(df)
                return f"CSV File Processed:\n{md_table}"
            except Exception as e:
                logger.error(f"Failed to process file {document_path}: {e}")
                return f"Failed to process file {document_path}: {e}"

        if document_path.endswith("xls"):
            output_path = document_path.replace(".xls", ".xlsx")
            x2x = XLS2XLSX(document_path)
            x2x.to_xlsx(output_path)
            document_path = output_path

        try:
            # Load the Excel workbook
            wb = load_workbook(document_path, data_only=True)
            sheet_info_list = []

            # Iterate through all sheets
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                if include_cell_info:
                    cell_info_list = []

                    for row in ws.iter_rows():
                        for cell in row:
                            # Skip cells that don't have proper coordinates (like
                            # merged cells)
                            if (
                                not hasattr(cell, 'column_letter')
                                or cell.value is None
                            ):
                                continue

                            row_num = cell.row
                            # Use getattr with fallback for column_letter
                            col_letter = getattr(cell, 'column_letter', 'A')

                            cell_value = cell.value

                            font_color = None
                            if (
                                cell.font
                                and cell.font.color
                                and "rgb=None" not in str(cell.font.color)
                            ):  # Handle font color
                                font_color = cell.font.color.rgb

                            fill_color = None
                            if (
                                cell.fill
                                and cell.fill.fgColor
                                and "rgb=None" not in str(cell.fill.fgColor)
                            ):  # Handle fill color
                                fill_color = cell.fill.fgColor.rgb

                            cell_info_list.append(
                                {
                                    "index": f"{row_num}{col_letter}",
                                    "value": cell_value,
                                    "font_color": font_color,
                                    "fill_color": fill_color,
                                }
                            )

                # Convert the sheet to a DataFrame and then to markdown
                sheet_df = pd.read_excel(
                    document_path, sheet_name=sheet, engine='openpyxl'
                )
                markdown_content = self._convert_to_markdown(sheet_df)

                # Collect all information for the sheet
                sheet_info = {
                    "sheet_name": sheet,
                    "markdown_content": markdown_content,
                }
                if include_cell_info:
                    sheet_info["cell_info_list"] = cell_info_list
                sheet_info_list.append(sheet_info)

            result_str = ""
            for sheet_info in sheet_info_list:
                result_str += f"\nSheet Name: {sheet_info['sheet_name']}\n"

                if include_cell_info:
                    result_str += f"""
            Cell information list:
            {sheet_info['cell_info_list']}
            """

                result_str += f"""
            Markdown View of the content:
            {sheet_info['markdown_content']}

            {'-' * 40}
            """

            return result_str
        except Exception as e:
            logger.error(f"Failed to process Excel file {document_path}: {e}")
            return f"Failed to process Excel file {document_path}: {e}"
